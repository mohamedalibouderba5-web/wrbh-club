"""Seed club structure + import Excel WRBH historique."""
from __future__ import annotations

import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.database import Base, SessionLocal, engine
from app.core.security import hash_password
from app.models import (
    Announcement,
    Athlete,
    Category,
    Club,
    CoachPayroll,
    Discipline,
    FeeInstallment,
    InventoryItem,
    LedgerEntry,
    Registration,
    Season,
    Team,
    TeamCoach,
    TeamMembership,
    User,
    Venue,
)

settings = get_settings()
EXCEL = Path(__file__).resolve().parents[2] / "gestion description joueur WRHB 1.3 (1).xlsm"

MONTHS_AR = {
    7: ("septembre", "سبتمبر"),
    8: ("octobre", "أكتوبر"),
    9: ("novembre", "نوفمبر"),
    10: ("décembre", "ديسمبر"),
    11: ("janvier", "جانفي"),
    12: ("février", "فيفري"),
    13: ("mars", "مارس"),
    14: ("avril", "افريل"),
    15: ("mai", "ماي"),
    16: ("juin", "جوان"),
    17: ("juillet", "جويلية"),
}

# Excel parametre (saison historique fichier)
CATS_2526 = [
    ("U14", "U14", "تحت 14", 2011, 2012),
    ("U13", "U13", "تحت 13", 2013, 2014),
    ("U11", "U11", "تحت 11", 2015, 2016),
    ("U9", "U9", "تحت 9", 2017, 2018),
    ("U7", "U7", "تحت 7", 2019, 2020),
    ("U6", "U6", "تحت 6", 2021, 2022),
]

# Affiche inscriptions 2026/2027
CATS_2627 = [
    ("U14", "U14", "تحت 14", 2012, 2013),
    ("U13", "U13", "تحت 13", 2014, 2015),
    ("U11", "U11", "تحت 11", 2016, 2017),
    ("U9", "U9", "تحت 9", 2018, 2019),
    ("U7", "U7", "تحت 7", 2020, 2021),
    ("U5", "U5", "تحت 5", 2022, 2023),
]


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.fromisoformat(str(val).replace("Z", "")).date()
    except Exception:
        return None


def ensure_base(db: Session) -> tuple[Club, Season, Season, Discipline]:
    club = db.query(Club).first()
    if not club:
        club = Club(
            name=settings.club_name,
            name_ar=settings.club_name_ar,
            acronym=settings.club_acronym,
            phone=settings.club_phone,
            whatsapp="+213540344884",
            address="Hammadi — face école Saray Hussein, route stade 01 Novembre 1954",
            logo_path="/logo.png",
            primary_color="#1E3A8A",
            accent_color="#F5C518",
            facebook="الوداد الرياضي لبلدية حمادي",
            instagram="الوداد الرياضي لبلدية حمادي",
        )
        db.add(club)
        db.flush()

    disc = db.query(Discipline).first()
    if not disc:
        disc = Discipline(club_id=club.id, name="Football", name_ar="كرة القدم", code="FOOT")
        db.add(disc)
        db.flush()

    season_old = db.query(Season).filter(Season.name == "2025/2026").first()
    if not season_old:
        season_old = Season(
            club_id=club.id,
            name="2025/2026",
            starts_on=date(2025, 9, 1),
            ends_on=date(2026, 8, 31),
            is_current=False,
            registration_open=False,
        )
        db.add(season_old)
        db.flush()
        for code, name, name_ar, y1, y2 in CATS_2526:
            cat = Category(
                season_id=season_old.id,
                discipline_id=disc.id,
                code=code,
                name=name,
                name_ar=name_ar,
                birth_year_min=y1,
                birth_year_max=y2,
            )
            db.add(cat)
            db.flush()
            # U13 split like Excel coaches u13 1 / u13 2
            if code == "U13":
                db.add(Team(category_id=cat.id, name="U13 Groupe 1", name_ar="U13 مجموعة 1", code="u13 1"))
                db.add(Team(category_id=cat.id, name="U13 Groupe 2", name_ar="U13 مجموعة 2", code="u13 2"))
            else:
                db.add(Team(category_id=cat.id, name=code, name_ar=name_ar, code=code.lower()))

    season = db.query(Season).filter(Season.name == "2026/2027").first()
    if not season:
        season = Season(
            club_id=club.id,
            name="2026/2027",
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
            is_current=True,
            registration_open=True,
        )
        db.add(season)
        db.flush()
        for code, name, name_ar, y1, y2 in CATS_2627:
            cat = Category(
                season_id=season.id,
                discipline_id=disc.id,
                code=code,
                name=name,
                name_ar=name_ar,
                birth_year_min=y1,
                birth_year_max=y2,
            )
            db.add(cat)
            db.flush()
            # Groupes demandés par le club : U14G1/G2 … U7G1, U5G1
            if code in {"U14", "U13", "U11", "U9"}:
                db.add(Team(category_id=cat.id, name=f"{code} Groupe 1", name_ar=f"{code} مجموعة 1", code=f"{code}G1"))
                db.add(Team(category_id=cat.id, name=f"{code} Groupe 2", name_ar=f"{code} مجموعة 2", code=f"{code}G2"))
            else:
                db.add(Team(category_id=cat.id, name=f"{code} Groupe 1", name_ar=f"{code} مجموعة 1", code=f"{code}G1"))

    if not db.query(User).filter(User.email == settings.default_admin_email).first():
        db.add(
            User(
                email=settings.default_admin_email,
                full_name="Administrateur WRBH",
                full_name_ar="مدير النادي",
                role="admin",
                password_hash=hash_password(settings.default_admin_password),
                phone="0540344884",
            )
        )

    if not db.query(Venue).first():
        db.add(
            Venue(
                club_id=club.id,
                name="Stade 01 Novembre 1954 — Hammadi",
                address="Hammadi",
                kind="pitch",
            )
        )

    if not db.query(Announcement).first():
        db.add(
            Announcement(
                club_id=club.id,
                title="Inscriptions saison 2026/2027 ouvertes",
                title_ar="فتح التسجيلات لموسم 2026/2027",
                body="Les inscriptions démarrent le 22 juillet 2026 (Mazhoud Foot / application mobile).",
                body_ar="تبدأ التسجيلات يوم 22 جويلية 2026.",
                audience="all",
                published_at=datetime.now(timezone.utc),
                is_pinned=True,
            )
        )

    if not db.query(InventoryItem).first():
        for name, qty in [("Ballons taille 4", 20), ("Chasubles", 40), ("Plots", 30), ("Plots d'entraînement", 50)]:
            db.add(InventoryItem(club_id=club.id, name=name, quantity=qty, alert_threshold=5))

    db.commit()
    return club, season_old, season, disc


def import_excel(db: Session, season: Season):
    if not EXCEL.exists():
        print("Excel introuvable:", EXCEL)
        return

    wb = load_workbook(EXCEL, data_only=True, keep_vba=False)
    cats = {c.code.upper(): c for c in db.query(Category).filter(Category.season_id == season.id)}
    # map also to 2025/26 categories for historical import
    hist = db.query(Season).filter(Season.name == "2025/2026").first()
    if hist:
        cats = {c.code.upper(): c for c in db.query(Category).filter(Category.season_id == hist.id)}
        season = hist

    teams_by_code = {}
    for t in db.query(Team).all():
        if t.code:
            teams_by_code[t.code.lower()] = t
        cat = db.get(Category, t.category_id)
        if cat and cat.season_id == season.id:
            teams_by_code[cat.code.lower()] = t

    # Players
    ws = wb["registre joueur"]
    imported = 0
    for row in ws.iter_rows(min_row=6, values_only=True):
        num, name = row[0], row[1]
        if not name or not isinstance(num, (int, float)):
            continue
        if db.query(Athlete).filter(Athlete.legacy_number == int(num)).first():
            continue
        birth = parse_date(row[2])
        place = row[3]
        reg_date = parse_date(row[4])
        cat_code = str(row[5]).upper().strip() if row[5] else None
        fee = row[6] if isinstance(row[6], (int, float)) else None
        status = str(row[20]) if row[20] else "Active"

        athlete = Athlete(
            legacy_number=int(num),
            full_name=str(name).strip(),
            full_name_ar=str(name).strip(),
            birth_date=birth,
            birth_place=str(place).strip() if place else None,
            status=status if status in ("Active", "Abandonne") else "Active",
        )
        db.add(athlete)
        db.flush()

        cat = cats.get(cat_code) if cat_code else None
        if not cat and birth:
            cat = (
                db.query(Category)
                .filter(
                    Category.season_id == season.id,
                    Category.birth_year_min <= birth.year,
                    Category.birth_year_max >= birth.year,
                )
                .first()
            )

        reg = Registration(
            athlete_id=athlete.id,
            season_id=season.id,
            category_id=cat.id if cat else None,
            registered_on=reg_date,
            status="approved",
            source="import",
            subscription_fee=Decimal(str(fee)) if fee is not None else None,
        )
        db.add(reg)
        db.flush()

        if fee:
            db.add(
                FeeInstallment(
                    athlete_id=athlete.id,
                    season_id=season.id,
                    registration_id=reg.id,
                    label="inscription",
                    label_ar="حقوق الاشتراك",
                    due_date=reg_date,
                    amount=Decimal(str(fee)),
                    amount_paid=Decimal(str(fee)),
                    status="paid",
                )
            )

        # monthly columns
        for col, (label, label_ar) in MONTHS_AR.items():
            val = row[col] if col < len(row) else None
            if isinstance(val, (int, float)) and val:
                db.add(
                    FeeInstallment(
                        athlete_id=athlete.id,
                        season_id=season.id,
                        registration_id=reg.id,
                        label=label,
                        label_ar=label_ar,
                        amount=Decimal(str(val)),
                        amount_paid=Decimal(str(val)),
                        status="paid",
                    )
                )

        if cat:
            team = teams_by_code.get(cat.code.lower())
            if team:
                db.add(
                    TeamMembership(
                        team_id=team.id,
                        athlete_id=athlete.id,
                        season_id=season.id,
                        is_active=True,
                    )
                )
        imported += 1

    # Coaches
    ws = wb["Entraineur"]
    for row in ws.iter_rows(min_row=3, max_row=7, values_only=True):
        num, name = row[0], row[1]
        if not name or not isinstance(num, (int, float)):
            continue
        email = f"coach{int(num)}@wrbh.local"
        if db.query(User).filter(User.email == email).first():
            continue
        coach = User(
            email=email,
            full_name=str(name).strip(),
            full_name_ar=str(name).strip(),
            role="coach",
            password_hash=hash_password("coach123"),
            birth_date=parse_date(row[2]),
            birth_place=str(row[3]).strip() if row[3] else None,
        )
        db.add(coach)
        db.flush()

        team_code = str(row[5]).strip().lower() if row[5] else None
        team = teams_by_code.get(team_code) if team_code else None
        if team:
            db.add(TeamCoach(team_id=team.id, user_id=coach.id))

        droit = row[6] if isinstance(row[6], (int, float)) else None
        if droit:
            db.add(
                CoachPayroll(
                    user_id=coach.id,
                    season_id=season.id,
                    pay_type="forfait",
                    label="حقوق المدرب",
                    amount=Decimal(str(droit)),
                    status="paid",
                    paid_on=parse_date(row[4]),
                )
            )
        for col, (label, _) in MONTHS_AR.items():
            val = row[col] if col < len(row) else None
            if isinstance(val, (int, float)) and val:
                db.add(
                    CoachPayroll(
                        user_id=coach.id,
                        season_id=season.id,
                        pay_type="monthly",
                        label=label,
                        amount=Decimal(str(val)),
                        status="paid",
                    )
                )

    # Transport expenses
    if "النقل" in wb.sheetnames:
        club = db.query(Club).first()
        ws = wb["النقل"]
        for row in ws.iter_rows(min_row=4, max_row=23, values_only=True):
            # cols F=owner G=amount H=date I=place J=num (0-index 5..)
            owner, amount, d, place = row[5], row[6], row[7], row[8]
            if amount is None:
                continue
            # amount can be "4500+1500"
            total = 0
            if isinstance(amount, (int, float)):
                total = float(amount)
            else:
                parts = str(amount).replace(" ", "").split("+")
                for p in parts:
                    try:
                        total += float(p)
                    except ValueError:
                        pass
            if total <= 0:
                continue
            db.add(
                LedgerEntry(
                    club_id=club.id,
                    season_id=season.id,
                    entry_type="expense",
                    category="transport",
                    label=f"Transport {place or ''}".strip(),
                    amount=Decimal(str(total)),
                    entry_date=parse_date(d) or date(2025, 11, 1),
                    counterparty=str(owner) if owner else None,
                    place=str(place) if place else None,
                )
            )

    # Demo parent
    if not db.query(User).filter(User.email == "parent@wrbh.local").first():
        parent = User(
            email="parent@wrbh.local",
            phone="0555000000",
            full_name="Parent Démo",
            full_name_ar="ولي أمر تجريبي",
            role="parent",
            password_hash=hash_password("parent123"),
        )
        db.add(parent)
        db.flush()
        # link first 2 athletes
        from app.models import ParentChild

        for a in db.query(Athlete).limit(2):
            db.add(ParentChild(parent_id=parent.id, athlete_id=a.id))

    db.commit()
    print(f"Import OK — joueurs ajoutés: {imported}")


def main():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        club, season_old, season_new, _ = ensure_base(db)
        print("Club:", club.name, club.name_ar)
        import_excel(db, season_old)
        print("Admin:", settings.default_admin_email, "/", settings.default_admin_password)
        print("Parent démo: parent@wrbh.local / parent123")
        print("Coachs: coach1@wrbh.local … coach5@wrbh.local / coach123")
    finally:
        db.close()


if __name__ == "__main__":
    main()
